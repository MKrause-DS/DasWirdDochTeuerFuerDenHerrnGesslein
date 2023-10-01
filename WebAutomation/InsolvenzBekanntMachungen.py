from time import sleep
from datetime import date, datetime
import os
from shutil import move
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import openpyxl

def initialization():
    # Current directory
    current_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Archive directory
    archive_dir = os.path.join(current_dir, "Archive")
    
    # Ensure the Archive directory exists
    if not os.path.exists(archive_dir):
        os.makedirs(archive_dir)

    # Today's date in the format YYYY-MM-DD
    today_str = date.today().strftime('%Y-%m-%d')
    
    # List all files in the directory
    for filename in os.listdir(current_dir):
        # Check for Protocol Excel files that don't have the current date in the filename
        if "Protocol" in filename and filename.endswith(".xlsx") and today_str not in filename:
            # Move the file to the Archive folder
            move(os.path.join(current_dir, filename), os.path.join(archive_dir, filename))

    # Create the webdriver
    driver = webdriver.Chrome('./chromedriver.exe')
    driver.get('https://neu.insolvenzbekanntmachungen.de/ap/suche.jsf')

    return driver

def write_to_protocol(status, message):
    protocol_file = f"Protocol_{date.today()}.xlsx"
    
    print(protocol_file)
    # Load the workbook or create a new one if it doesn't exist
    try:
        wb = openpyxl.load_workbook(protocol_file)
        ws = wb.active
        
        # If there's no active worksheet, create a new one
        if ws is None:
            ws = wb.create_sheet("Sheet1")
            
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['TimeStamp', 'Status', 'Message'])
    
    # Append the current date & time, status, and message to the Excel file
    ws.append([datetime.now().strftime('%Y-%m-%d %H:%M:%S'), status, message])

    wb.save(protocol_file)


def search_company(CompanyName, driver):
    

    max_retries = 3
    error_counter = 0

    while error_counter < max_retries:
        try:
            # Attempt to find the firmenname element
            firmenname = driver.find_element(By.ID, 'frm_suche:ireg_registereintrag:itx_registernummer')

            # If successful, send the company name
            firmenname.send_keys(CompanyName)
            sleep(10)
            write_to_protocol('Successfull','Company ' + CompanyName +'found')
            break
        except:
            error_counter += 1
            print(f"Attempt {error_counter} failed. Retrying...")
            write_to_protocol('Retry','Text')

    if error_counter == max_retries:
        print("Failed to find the element after maximum retries. Exiting.")
        write_to_protocol('Failed','Text')
        driver.quit()
        return 

if __name__ == "__main__":

    
    driver=initialization()

    search_company('asd',driver)

    driver.quit()
    